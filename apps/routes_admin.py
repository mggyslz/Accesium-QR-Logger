
from flask import Blueprint, render_template, request, jsonify, url_for, session, redirect, send_file
from core.auth_decorators import admin_login_required, add_no_cache_headers
import re 
from core.database import (
    get_all_users, delete_user,
    get_admin_by_username_or_email,
    get_total_inside, get_recent_logs,
    get_daily_counts, get_hourly_counts,
    get_current_inside, export_logs_csv, get_conn,
    utc_to_ph_time,
    get_active_location, set_active_location, PREDEFINED_LOCATIONS,
    get_user_by_id, get_user_email, add_admin_with_email
)
from core.trusted_device_utils import get_user_trusted_devices
from core.validation import (
    validate_user_registration,
    validate_username,
    validate_email,
    validate_name,
    validate_role,
    sanitize_input,
    csrf_protect,
    rate_limit,
    generate_csrf_token,
    get_client_ip,
    log_suspicious_activity,
    MAX_USERNAME_LENGTH,
    MAX_EMAIL_LENGTH,
    MAX_NAME_LENGTH
)
from datetime import datetime
from pathlib import Path
import pandas as pd
from io import BytesIO

admin_bp = Blueprint('admin', __name__, template_folder='../templates', url_prefix='/admin')

@admin_bp.context_processor
def inject_csrf_token():
    return dict(csrf_token=generate_csrf_token)

def generate_username_from_name(full_name):
    """
    Generate username from full name in firstname.lastname format
    Handles edge cases like single names, multiple names, special characters
    """
    import re
    from core.database import get_conn
    name_parts = full_name.strip().lower().split()
    
    if not name_parts:
        raise ValueError("Name cannot be empty")
    name_parts = [re.sub(r'[^a-z0-9]', '', part) for part in name_parts]
    name_parts = [part for part in name_parts if part]  
    if not name_parts:
        raise ValueError("Name contains no valid characters")
    if len(name_parts) == 1:   
        base_username = name_parts[0]
    else: 
        base_username = f"{name_parts[0]}.{name_parts[-1]}"
    if len(base_username) < 3:
        raise ValueError("Generated username is too short. Please provide a longer name.")
    username = base_username
    counter = 1
    conn = get_conn()
    cur = conn.cursor()
    while True:
        cur.execute("SELECT COUNT(*) FROM users WHERE username = ?", (username,))
        if cur.fetchone()[0] == 0:
            break
        username = f"{base_username}{counter}"
        counter += 1
    conn.close()
    return username

@admin_bp.route('/')
@admin_login_required 
def dashboard():
    users = get_all_users()
    total_inside = get_total_inside()
    recent_logs = get_recent_logs(limit=10)
    daily_counts = get_daily_counts(days=7)
    hourly_counts = get_hourly_counts()
    current_inside = get_current_inside()
    active_location = get_active_location()
    locations = PREDEFINED_LOCATIONS
    devices = get_user_trusted_devices('admin', session['admin_id'])
    
    from core.database import is_first_admin
    is_first_admin_flag = is_first_admin(session['admin_id'])
    
    # Get today's scan count
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT COUNT(*) FROM access_logs 
        WHERE DATE(timestamp) = DATE('now')
    """)
    today_scans = cur.fetchone()[0]
    conn.close()
    
    chart_labels = [row[0] for row in daily_counts]
    chart_ins = [row[1] for row in daily_counts]
    chart_outs = [row[2] for row in daily_counts]
    hourly_labels = [f"{h}:00" for h, _, _ in hourly_counts]
    hourly_ins = [row[1] for row in hourly_counts]
    hourly_outs = [row[2] for row in hourly_counts]

    return render_template(
        'admin.html',
        users=users,
        total_inside=total_inside,
        today_scans=today_scans,
        recent_logs=recent_logs,
        chart_labels=chart_labels,
        chart_ins=chart_ins,
        chart_outs=chart_outs,
        hourly_labels=hourly_labels,
        hourly_ins=hourly_ins,
        hourly_outs=hourly_outs,
        current_inside=current_inside,
        active_location=active_location,
        locations=locations,
        devices=devices,
        is_first_admin=is_first_admin_flag 
    )

@admin_bp.route('/set-location', methods=['POST'])
@csrf_protect
@admin_login_required 
def set_location():
    """Set the active location for logging."""
    location = sanitize_input(request.form.get('location', ''), 50)
    if not location:
        return jsonify({"status": "error", "message": "Location is required"}), 400
    try:
        set_active_location(location)
        return jsonify({"status": "success", "message": f"Active location set to {location}"})
    except ValueError as ve:
        return jsonify({"status": "error", "message": str(ve)}), 400
    except Exception as e:
        log_suspicious_activity('set_location_error', {'error': str(e), 'location': location})
        return jsonify({"status": "error", "message": "Failed to set location"}), 500

@admin_bp.route('/delete/<int:user_id>', methods=['POST'])
@csrf_protect
@admin_login_required 
def delete_user_route(user_id):
    try:
        delete_user(user_id)
        return jsonify({"status": "success", "message": "User deleted"})
    except Exception as e:
        log_suspicious_activity('delete_user_error', {'user_id': user_id, 'error': str(e)})
        return jsonify({"status": "error", "message": "Failed to delete user"}), 500


@admin_bp.route('/stats', methods=['GET'])
@admin_login_required 
def stats():
    try:
        total_inside = get_total_inside()
        recent_logs = get_recent_logs(limit=10)
        daily_counts = get_daily_counts(days=7)
        hourly_counts = get_hourly_counts()
        current_inside = get_current_inside()
        active_location = get_active_location()
        
        # Get today's scan count
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM access_logs 
            WHERE DATE(timestamp) = DATE('now')
        """)
        today_scans = cur.fetchone()[0]
        conn.close()
        
        return jsonify({
            "status": "success",
            "total_inside": total_inside,
            "today_scans": today_scans,
            "recent_logs": [
                {"log_id": r[0], "name": r[1], "action": r[2], "timestamp": r[3], "location": r[4]}
                for r in recent_logs
            ],
            "daily_counts": [{"day": d[0], "ins": d[1], "outs": d[2]} for d in daily_counts],
            "hourly_counts": [{"hour": h[0], "ins": h[1], "outs": h[2]} for h in hourly_counts],
            "current_inside": [
                {"user_id": u[0], "name": u[1], "role": u[2], "timestamp": u[3]}
                for u in current_inside
            ],
            "active_location": active_location
        })
    except Exception as e:
        log_suspicious_activity('stats_error', {'error': str(e)})
        return jsonify({"status": "error", "message": "Failed to retrieve stats"}), 500


@admin_bp.route('/export-logs', methods=['GET']) 
@admin_login_required 
def export_logs():
    """Export access logs to timestamped CSV and return as download."""
    if 'admin_id' not in session:
        return jsonify({"status": "error", "message": "Not authorized"}), 401
    try:
        exports_dir = Path('static/exports')
        exports_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'access_logs_{timestamp}.csv'
        filepath = exports_dir / filename
        export_logs_csv(str(filepath))
        return send_file(
            filepath,
            as_attachment=True,
            download_name=filename,
            mimetype='text/csv'
        )
    except FileNotFoundError:
        return jsonify({"status": "error", "message": "Export file not found"}), 404
    except Exception as e:
        log_suspicious_activity('export_logs_error', {'error': str(e)})
        return jsonify({"status": "error", "message": "Export failed"}), 500


@admin_bp.route('/attendance', methods=['GET']) 
@admin_login_required
def get_attendance():
    """Get attendance records with pagination and filters"""
    if 'admin_id' not in session:
        return jsonify({"status": "error", "message": "Not authorized"}), 401
    try:
        try:
            page = max(1, int(request.args.get('page', 1)))
            per_page = min(200, max(10, int(request.args.get('per_page', 10))))
        except (ValueError, TypeError):
            page = 1
            per_page = 10

        date_from = sanitize_input(request.args.get('date_from', ''), 10)
        date_to = sanitize_input(request.args.get('date_to', ''), 10)
        search = sanitize_input(request.args.get('search', ''), 100)
        action_filter = sanitize_input(request.args.get('action', ''), 10)
        location_filter = sanitize_input(request.args.get('location', ''), 50)

        if date_from and not re.match(r'^\d{4}-\d{2}-\d{2}$', date_from):
            return jsonify({"status": "error", "message": "Invalid date_from format"}), 400
        if date_to and not re.match(r'^\d{4}-\d{2}-\d{2}$', date_to):
            return jsonify({"status": "error", "message": "Invalid date_to format"}), 400
        if action_filter and action_filter not in ['IN', 'OUT']:
            return jsonify({"status": "error", "message": "Invalid action filter"}), 400
        conn = get_conn()
        cur = conn.cursor()
        query = """
            SELECT l.log_id, u.name, u.role, l.action, l.timestamp, l.location
            FROM access_logs l
            JOIN users u ON l.user_id = u.user_id
            WHERE 1=1
        """
        params = []

        if date_from:
            query += " AND DATE(l.timestamp) >= ?"
            params.append(date_from)
        
        if date_to:
            query += " AND DATE(l.timestamp) <= ?"
            params.append(date_to)
        
        if search:
            query += " AND (u.name LIKE ? OR u.role LIKE ?)"
            params.extend([f'%{search}%', f'%{search}%'])
        
        if action_filter:
            query += " AND l.action = ?"
            params.append(action_filter)
        
        if location_filter:
            query += " AND l.location = ?"
            params.append(location_filter)
            
        count_query = f"SELECT COUNT(*) FROM ({query})"
        cur.execute(count_query, params)
        total_records = cur.fetchone()[0]
        total_pages = (total_records + per_page - 1) // per_page
        query += " ORDER BY l.timestamp DESC LIMIT ? OFFSET ?"
        params.extend([per_page, (page - 1) * per_page])
        cur.execute(query, params)
        records = cur.fetchall()
        cur.execute("SELECT DISTINCT location FROM access_logs WHERE location IS NOT NULL ORDER BY location")
        locations = [row[0] for row in cur.fetchall()]
        conn.close()
        records_list = [
            {
                'log_id': r[0],
                'name': r[1],
                'role': r[2],
                'action': r[3],
                'timestamp': utc_to_ph_time(r[4]),
                'location': r[5] or 'Gate'
            }
            for r in records
        ]

        return jsonify({
            "status": "success",
            "records": records_list,
            "total_records": total_records,
            "total_pages": total_pages,
            "current_page": page,
            "locations": locations
        })

    except Exception as e:
        log_suspicious_activity('attendance_error', {'error': str(e)})
        return jsonify({"status": "error", "message": "Failed to retrieve attendance"}), 500

@admin_bp.route('/export-attendance-excel', methods=['GET'])
@admin_login_required
def export_attendance_excel():
    """Export attendance records to Excel with filters"""
    if 'admin_id' not in session:
        return jsonify({"status": "error", "message": "Not authorized"}), 401
    try:
        date_from = sanitize_input(request.args.get('date_from', ''), 10)
        date_to = sanitize_input(request.args.get('date_to', ''), 10)
        search = sanitize_input(request.args.get('search', ''), 100)
        action_filter = sanitize_input(request.args.get('action', ''), 10)
        location_filter = sanitize_input(request.args.get('location', ''), 50)
        if date_from and not re.match(r'^\d{4}-\d{2}-\d{2}$', date_from):
            return jsonify({"status": "error", "message": "Invalid date format"}), 400
        
        if date_to and not re.match(r'^\d{4}-\d{2}-\d{2}$', date_to):
            return jsonify({"status": "error", "message": "Invalid date format"}), 400

        if action_filter and action_filter not in ['IN', 'OUT', '']:
            return jsonify({"status": "error", "message": "Invalid action filter"}), 400
        conn = get_conn()
        
        query = """
            SELECT 
                u.name AS 'Name',
                u.role AS 'Role',
                l.action AS 'Action',
                l.timestamp AS 'Timestamp',
                l.location AS 'Location'
            FROM access_logs l
            JOIN users u ON l.user_id = u.user_id
            WHERE 1=1
        """

        params = []
        if date_from:
            query += " AND DATE(l.timestamp) >= ?"
            params.append(date_from)
        
        if date_to:
            query += " AND DATE(l.timestamp) <= ?"
            params.append(date_to)
        
        if search:
            query += " AND (u.name LIKE ? OR u.role LIKE ?)"
            params.extend([f'%{search}%', f'%{search}%'])
        
        if action_filter:
            query += " AND l.action = ?"
            params.append(action_filter)
        
        if location_filter:
            query += " AND l.location = ?"
            params.append(location_filter)

        query += " ORDER BY l.timestamp DESC"

        df = pd.read_sql_query(query, conn, params=params)
        conn.close()

        if df.empty:
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                empty_df = pd.DataFrame({'Message': ['No records found matching your criteria']})
                empty_df.to_excel(writer, sheet_name='No Data', index=False)
            
            output.seek(0)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'attendance_{timestamp}.xlsx'
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        if 'Timestamp' in df.columns:
            df['Timestamp'] = df['Timestamp'].apply(lambda x: utc_to_ph_time(x) if pd.notna(x) else '')
            df[['Date', 'Time']] = df['Timestamp'].str.split(' ', n=1, expand=True)
            df = df[['Name', 'Role', 'Action', 'Date', 'Time', 'Location']]
        df['Location'] = df['Location'].fillna('Gate')

        df_in = df[df['Action'] == 'IN'].copy()
        df_out = df[df['Action'] == 'OUT'].copy()
        df_in_export = df_in.drop('Action', axis=1) if 'Action' in df_in.columns else df_in
        df_out_export = df_out.drop('Action', axis=1) if 'Action' in df_out.columns else df_out
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            from openpyxl.styles import Font, PatternFill, Alignment

            def style_sheet(worksheet, df_data, header_color="366092"):
                if df_data.empty:
                    return
                
                header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                for idx, col in enumerate(df_data.columns):
                    max_length = max(
                        df_data[col].astype(str).str.len().max() if len(df_data) > 0 else 0,
                        len(str(col))
                    ) + 2
                    col_letter = chr(65 + idx) if idx < 26 else chr(65 + idx // 26 - 1) + chr(65 + idx % 26)
                    worksheet.column_dimensions[col_letter].width = min(max_length, 50)
                
                for row in worksheet.iter_rows(min_row=2, max_row=len(df_data) + 1):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            df.to_excel(writer, sheet_name='All Records', index=False)
            style_sheet(writer.sheets['All Records'], df)
            if not df_in_export.empty:
                df_in_export.to_excel(writer, sheet_name='IN Records', index=False)
                style_sheet(writer.sheets['IN Records'], df_in_export, header_color="28a745")
            if not df_out_export.empty:
                df_out_export.to_excel(writer, sheet_name='OUT Records', index=False)
                style_sheet(writer.sheets['OUT Records'], df_out_export, header_color="e74c3c")
            summary_df = pd.DataFrame({
                'Metric': ['Total Records', 'Total IN', 'Total OUT', 'Unique Users', 'Unique Locations', 'Date Range', 'Export Date'],
                'Value': [
                    len(df), len(df_in), len(df_out),
                    df['Name'].nunique() if 'Name' in df.columns else 0,
                    df['Location'].nunique() if 'Location' in df.columns else 0,
                    f"{date_from or 'All'} to {date_to or 'All'}",
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
            })
            
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            style_sheet(writer.sheets['Summary'], summary_df)
        output.seek(0)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'attendance_{timestamp}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        log_suspicious_activity('export_excel_error', {'error': str(e)})
        return jsonify({"status": "error", "message": "Export failed"}), 500


@admin_bp.route('/add-user', methods=['POST'])
@csrf_protect
@admin_login_required
def add_user_by_admin_route():
    """Admin creates a new user with password"""
    if 'admin_id' not in session:
        return jsonify({"status": "error", "message": "Not authorized"}), 401
    email = sanitize_input(request.form.get('email', ''), MAX_EMAIL_LENGTH)
    name = sanitize_input(request.form.get('name', ''), MAX_NAME_LENGTH)
    role = sanitize_input(request.form.get('role', 'Student'), 50)
    username = generate_username_from_name(name)
    valid, error = validate_username(username)
    if not valid:
        return jsonify({"status": "error", "message": f"Generated username invalid: {error}"}), 400

    valid, error = validate_user_registration(
        username=username,
        email=email,
        name=name,
        password='TempPass123!',  
        role=role
    )
    
    if not valid:
        log_suspicious_activity('invalid_user_registration', {
            'error': error,
            'username': username,
            'email': email
        })
        return jsonify({"status": "error", "message": error}), 400
    try:
        from core.email_utils import generate_temporary_password, send_new_user_credentials_email, is_smtp_configured
        from core.database import add_user_by_admin
        from core.notification_utils import notify_new_user_registered
        
        if not is_smtp_configured():
            return jsonify({
                "status": "error", 
                "message": "Email service not configured. Cannot create user account."
            }), 500
        password = generate_temporary_password()
        user_id, qr_token, qr_filename, pin = add_user_by_admin(
            username=username,
            email=email,
            name=name,
            role=role,
            password=password
        )
        login_url = url_for('user.login_page', _external=True)
        email_sent = send_new_user_credentials_email(
            email=email,
            name=name,
            username=username,
            password=password,
            pin=pin,  
            login_url=login_url
        )
        if not email_sent:
            from core.database import delete_user
            delete_user(user_id)
            return jsonify({
                "status": "error",
                "message": "Failed to send credentials email. User creation cancelled."
            }), 500

        notify_new_user_registered(name, role)

        return jsonify({
            "status": "success",
            "message": f"User '{name}' created successfully with username '{username}'. Credentials sent to {email}.",
            "qr_path": url_for('static', filename=f'qrcodes/{qr_filename}'),
            "username": username,
            "pin": pin  
        })

    except ValueError as ve:
        log_suspicious_activity('user_creation_failed', {'error': str(ve), 'username': username})
        return jsonify({"status": "error", "message": str(ve)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        log_suspicious_activity('user_creation_error', {'error': str(e), 'username': username})
        return jsonify({"status": "error", "message": "Error creating user"}), 500
@admin_bp.route('/change-password', methods=['POST'])
@csrf_protect
@admin_login_required
def change_password():
    """Admin changes their own password"""
    if 'admin_id' not in session:
        return jsonify({"status": "error", "message": "Not authorized"}), 401
    
    try:
        from core.security import verify_pin, generate_salt, hash_pin
        from core.email_utils import send_password_changed_notification_email
        
        admin_id = session['admin_id']
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        if not current_password or not new_password or not confirm_password:
            return jsonify({"status": "error", "message": "All fields are required"}), 400
        
        if new_password != confirm_password:
            return jsonify({"status": "error", "message": "New passwords do not match"}), 400
        from core.validation import validate_password
        valid, error = validate_password(new_password)
        if not valid:
            return jsonify({"status": "error", "message": error}), 400
        admin = get_admin_by_username_or_email(session.get('admin_username', ''))
        if not admin:
            return jsonify({"status": "error", "message": "Admin not found"}), 404
        admin_id_db, username, email, name, pass_hash, pass_salt, email_verified = admin
        if not verify_pin(current_password, pass_salt, pass_hash):
            log_suspicious_activity('admin_wrong_current_password', {
                'admin_id': admin_id,
                'username': username
            })
            return jsonify({"status": "error", "message": "Current password is incorrect"}), 401
        new_salt = generate_salt()
        new_hash = hash_pin(new_password, new_salt)
        conn = get_conn()
        try:
            cur = conn.cursor()
            cur.execute(
                "UPDATE admins SET pass_hash = ?, pass_salt = ? WHERE admin_id = ?",
                (new_hash, new_salt, admin_id)
            )
            conn.commit()
        finally:
            conn.close()
        try:
            if email:
                send_password_changed_notification_email(email, name or username)
        except Exception as e:
            log_suspicious_activity('password_change_email_error', {
                'admin_id': admin_id,
                'error': str(e)
            })
        return jsonify({
            "status": "success",
            "message": "Password changed successfully"
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        log_suspicious_activity('admin_password_change_error', {
            'admin_id': session.get('admin_id'),
            'error': str(e)
        })
        return jsonify({"status": "error", "message": "Failed to change password"}), 500

@admin_bp.route('/access-rules')
@admin_login_required
def access_rules_page():
    """View all access rules"""
    from core.access_control import get_all_access_rules_for_admin
    rules = get_all_access_rules_for_admin()
    users = get_all_users()
    
    return render_template(
        'admin.html',
        access_rules=rules,
        users_for_rules=users
    )


@admin_bp.route('/access-rules/add', methods=['POST'])
@csrf_protect
@admin_login_required
def add_access_rule():
    """Add new access rule"""
    from core.access_control import add_access_rule
    
    try:
        user_id = int(request.form.get('user_id'))
        rule_type = sanitize_input(request.form.get('rule_type'), 20)
        location = sanitize_input(request.form.get('location', ''), 50) or None
        time_from = sanitize_input(request.form.get('time_from', ''), 10) or None
        time_to = sanitize_input(request.form.get('time_to', ''), 10) or None
        date_from = sanitize_input(request.form.get('date_from', ''), 10) or None
        date_to = sanitize_input(request.form.get('date_to', ''), 10) or None
        specific_dates = sanitize_input(request.form.get('specific_dates', ''), 200) or None
        
        if rule_type not in ['whitelist', 'blacklist']:
            return jsonify({"status": "error", "message": "Invalid rule type"}), 400
        
        rule_id = add_access_rule(
            user_id, rule_type, location, time_from, time_to, 
            date_from, date_to, specific_dates
        )
        
        return jsonify({
            "status": "success",
            "message": f"{rule_type.capitalize()} rule added successfully",
            "rule_id": rule_id
        })
    except Exception as e:
        log_suspicious_activity('add_access_rule_error', {'error': str(e)})
        return jsonify({"status": "error", "message": "Failed to add rule"}), 500


@admin_bp.route('/access-rules/delete/<int:rule_id>', methods=['POST'])
@csrf_protect
@admin_login_required
def delete_access_rule_route(rule_id):
    """Delete access rule"""
    from core.access_control import delete_access_rule
    
    try:
        delete_access_rule(rule_id)
        return jsonify({"status": "success", "message": "Rule deleted"})
    except Exception as e:
        log_suspicious_activity('delete_access_rule_error', {'error': str(e)})
        return jsonify({"status": "error", "message": "Failed to delete rule"}), 500


@admin_bp.route('/access-rules/toggle/<int:rule_id>', methods=['POST'])
@csrf_protect
@admin_login_required
def toggle_access_rule(rule_id):
    """Toggle rule enabled/disabled"""
    from core.access_control import toggle_rule_enabled
    
    try:
        enabled = request.form.get('enabled') == '1'
        toggle_rule_enabled(rule_id, enabled)
        return jsonify({"status": "success", "message": "Rule updated"})
    except Exception as e:
        log_suspicious_activity('toggle_access_rule_error', {'error': str(e)})
        return jsonify({"status": "error", "message": "Failed to update rule"}), 500


@admin_bp.route('/access-rules/data', methods=['GET'])
@admin_login_required
def get_access_rules_data():
    """Get all access rules as JSON"""
    from core.access_control import get_all_access_rules_for_admin
    
    try:
        rules = get_all_access_rules_for_admin()
        return jsonify({
            "status": "success",
            "rules": [
                {
                    "rule_id": r[0],
                    "user_name": r[1],
                    "rule_type": r[2],
                    "location": r[3],
                    "time_from": r[4],
                    "time_to": r[5],
                    "date_from": r[6],
                    "date_to": r[7],
                    "specific_dates": r[8],
                    "enabled": r[9],
                    "created_at": r[10]
                } for r in rules
            ]
        })
    except Exception as e:
        log_suspicious_activity('get_access_rules_error', {'error': str(e)})
        return jsonify({"status": "error", "message": "Failed to load rules"}), 500

@admin_bp.route('/search-users', methods=['GET'])
@admin_login_required
def search_users():
    """Search and filter users for access rules"""
    try:
        search_query = sanitize_input(request.args.get('search', ''), 100).lower()
        roles_filter = request.args.getlist('roles[]')  
        
        conn = get_conn()
        cur = conn.cursor()
        
        query = "SELECT user_id, username, name, role, email FROM users WHERE 1=1"
        params = []
        
        
        if search_query:
            query += """ AND (
                LOWER(name) LIKE ? OR 
                LOWER(username) LIKE ? OR 
                LOWER(email) LIKE ? OR 
                CAST(user_id AS TEXT) LIKE ?
            )"""
            search_param = f'%{search_query}%'
            params.extend([search_param, search_param, search_param, search_param])
        
        
        if roles_filter:
            placeholders = ','.join('?' * len(roles_filter))
            query += f" AND role IN ({placeholders})"
            params.extend(roles_filter)
        
        query += " ORDER BY name ASC"
        
        cur.execute(query, params)
        users = cur.fetchall()
        conn.close()
        
        return jsonify({
            "status": "success",
            "users": [
                {
                    "user_id": u[0],
                    "username": u[1],
                    "name": u[2],
                    "role": u[3],
                    "email": u[4]
                } for u in users
            ]
        })
    except Exception as e:
        log_suspicious_activity('search_users_error', {'error': str(e)})
        return jsonify({"status": "error", "message": "Failed to search users"}), 500


@admin_bp.route('/bulk-add-access-rules', methods=['POST'])
@csrf_protect
@admin_login_required
def bulk_add_access_rules():
    """Add access rules to multiple users at once"""
    from core.access_control import add_access_rule
    
    try:
        user_ids = request.form.getlist('user_ids[]')
        rule_type = sanitize_input(request.form.get('rule_type'), 20)
        location = sanitize_input(request.form.get('location', ''), 50) or None
        time_from = sanitize_input(request.form.get('time_from', ''), 10) or None
        time_to = sanitize_input(request.form.get('time_to', ''), 10) or None
        date_from = sanitize_input(request.form.get('date_from', ''), 10) or None
        date_to = sanitize_input(request.form.get('date_to', ''), 10) or None
        specific_dates = sanitize_input(request.form.get('specific_dates', ''), 200) or None
        
        if not user_ids:
            return jsonify({"status": "error", "message": "No users selected"}), 400
        if rule_type not in ['whitelist', 'blacklist']:
            return jsonify({"status": "error", "message": "Invalid rule type"}), 400
        added_count = 0
        for user_id in user_ids:
            try:
                add_access_rule(
                    int(user_id), rule_type, location, time_from, time_to,
                    date_from, date_to, specific_dates
                )
                added_count += 1
            except Exception as e:
                print(f"Failed to add rule for user {user_id}: {str(e)}")
        
        return jsonify({
            "status": "success",
            "message": f"Added {added_count} rule(s) to {len(user_ids)} user(s)",
            "added_count": added_count
        })
    except Exception as e:
        log_suspicious_activity('bulk_add_rules_error', {'error': str(e)})
        return jsonify({"status": "error", "message": "Failed to add bulk rules"}), 500
    
@admin_bp.route('/update-session-settings', methods=['POST'])
@csrf_protect
@admin_login_required
def update_session_settings():
    """Update session timeout settings"""
    if 'admin_id' not in session:
        return jsonify({"status": "error", "message": "Not authorized"}), 401
    try:
        from core.settings_manager import update_session_timeout
        timeout = int(request.form.get('session_timeout', 30))
        
        if not 5 <= timeout <= 480:
            return jsonify({
                "status": "error",
                "message": "Session timeout must be between 5 and 480 minutes"
            }), 400
        
        if update_session_timeout(timeout, session['admin_id']):
            return jsonify({
                "status": "success",
                "message": f"Session timeout updated to {timeout} minutes"
            })
        else:
            return jsonify({
                "status": "error",
                "message": "Failed to update session timeout"
            }), 500
            
    except ValueError:
        return jsonify({
            "status": "error",
            "message": "Invalid timeout value"
        }), 400
    except Exception as e:
        log_suspicious_activity('update_session_settings_error', {
            'error': str(e),
            'admin_id': session['admin_id']
        })
        return jsonify({
            "status": "error",
            "message": "Failed to update settings"
        }), 500


@admin_bp.route('/update-login-security', methods=['POST'])
@csrf_protect
@admin_login_required
def update_login_security():
    """Update login security settings"""
    if 'admin_id' not in session:
        return jsonify({"status": "error", "message": "Not authorized"}), 401

    try:
        from core.settings_manager import (
            update_2fa_requirement,
            update_lockout_policy,
            update_lockout_threshold
        )

        require_2fa = request.form.get('enable2FA') == '1'
        lockout_enabled = request.form.get('lockoutPolicy') == '1'

        # -------------------------------
        # OPTIONAL lockout threshold logic
        # -------------------------------
        lockout_threshold_str = request.form.get('lockoutThreshold', '5')

        try:
            lockout_threshold = int(lockout_threshold_str) if lockout_threshold_str else 5
        except (ValueError, TypeError):
            lockout_threshold = 5

        # Out of range → fallback to default, no error
        if not 3 <= lockout_threshold <= 10:
            lockout_threshold = 5

        admin_id = session['admin_id']

        update_2fa_requirement(require_2fa, admin_id)
        update_lockout_policy(lockout_enabled, admin_id)
        update_lockout_threshold(lockout_threshold, admin_id)

        return jsonify({
            "status": "success",
            "message": "Security settings updated successfully"
        })

    except Exception as e:
        log_suspicious_activity('update_login_security_error', {
            'error': str(e),
            'admin_id': session['admin_id']
        })
        return jsonify({
            "status": "error",
            "message": "Failed to update security settings"
        }), 500



@admin_bp.route('/get-system-settings', methods=['GET'])
@admin_login_required
def get_system_settings():
    """Get current system settings"""
    if 'admin_id' not in session:
        return jsonify({"status": "error", "message": "Not authorized"}), 401
    
    try:
        from core.settings_manager import SystemSettings
        
        settings = {
            'session_timeout': SystemSettings.get_session_timeout(),
            'require_2fa': SystemSettings.is_2fa_required(),
            'lockout_enabled': SystemSettings.is_lockout_enabled(),
            'lockout_threshold': SystemSettings.get_lockout_threshold(),
            'lockout_duration': SystemSettings.get_lockout_duration()
        }
        
        return jsonify({
            "status": "success",
            "settings": settings
        })
    except Exception as e:
        log_suspicious_activity('get_system_settings_error', {
            'error': str(e)
        })
        return jsonify({
            "status": "error",
            "message": "Failed to retrieve settings"
        }), 500

@admin_bp.route('/reset-user-password', methods=['POST'])
@csrf_protect
@admin_login_required
def reset_user_password():
    """
    Admin resets a user's password
    Generates a new temporary password and emails it to the user
    """
    if 'admin_id' not in session:
        return jsonify({"status": "error", "message": "Not authorized"}), 401
    
    try:
        from core.email_utils import (
            generate_temporary_password, 
            send_password_reset_by_admin_email,
            is_smtp_configured
        )
        from core.security import generate_salt, hash_pin
        from core.notification_utils import notify_password_changed
        
        
        user_id = request.form.get('user_id')
        if not user_id:
            return jsonify({"status": "error", "message": "User ID is required"}), 400
        
        try:
            user_id = int(user_id)
        except (ValueError, TypeError):
            return jsonify({"status": "error", "message": "Invalid user ID"}), 400
        
        
        if not is_smtp_configured():
            return jsonify({
                "status": "error",
                "message": "Email service not configured. Cannot send password reset."
            }), 500
        
        
        user = get_user_by_id(user_id)
        if not user:
            return jsonify({"status": "error", "message": "User not found"}), 404
        
        
        user_id_db, username, name, role, qr_code, qr_token, status = user
        
        
        user_email = get_user_email(user_id)
        if not user_email:
            return jsonify({
                "status": "error",
                "message": "User does not have an email address registered. Cannot send password reset."
            }), 400
        temporary_password = generate_temporary_password()
        password_salt = generate_salt()
        password_hash = hash_pin(temporary_password, password_salt)
        conn = get_conn()
        try:
            cur = conn.cursor()
            cur.execute("""
                UPDATE users 
                SET password_hash = ?, password_salt = ?, force_password_change = 1
                WHERE user_id = ?
            """, (password_hash, password_salt, user_id))
            conn.commit()
        finally:
            conn.close()
        login_url = url_for('user.login_page', _external=True)
        email_sent = send_password_reset_by_admin_email(
            email=user_email,
            name=name,
            temporary_password=temporary_password,
            login_url=login_url
        )
        if not email_sent:
            
            conn = get_conn()
            try:
                cur = conn.cursor()
                cur.execute("""
                    UPDATE users 
                    SET force_password_change = 0
                    WHERE user_id = ?
                """, (user_id,))
                conn.commit()
            finally:
                conn.close()
            
            return jsonify({
                "status": "error",
                "message": "Failed to send password reset email. Password was not changed."
            }), 500
        notify_password_changed('user', user_id)   
        log_suspicious_activity('admin_reset_user_password', {
            'admin_id': session['admin_id'],
            'admin_username': session.get('admin_username'),
            'target_user_id': user_id,
            'target_username': username,
            'target_email': user_email
        })
        return jsonify({
            "status": "success",
            "message": f"Password reset successful. A temporary password has been sent to {user_email}."
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        log_suspicious_activity('reset_user_password_error', {
            'admin_id': session.get('admin_id'),
            'error': str(e)
        })
        return jsonify({
            "status": "error",
            "message": "Failed to reset password. Please try again."
        }), 500
        
@admin_bp.route('/add-admin', methods=['POST'])
@csrf_protect
@admin_login_required
def add_admin_route():
    """Add new admin account (only first admin can do this)"""
    if 'admin_id' not in session:
        return jsonify({"status": "error", "message": "Not authorized"}), 401
    
    
    from core.database import is_first_admin
    if not is_first_admin(session['admin_id']):
        log_suspicious_activity('unauthorized_admin_creation', {
            'attempted_by': session['admin_id'],
            'attempted_username': session.get('admin_username')
        })
        return jsonify({
            "status": "error", 
            "message": "Only the first administrator can create new admin accounts."
        }), 403
    username = sanitize_input(request.form.get('username', ''), MAX_USERNAME_LENGTH)
    email = sanitize_input(request.form.get('email', ''), MAX_EMAIL_LENGTH).lower()
    name = sanitize_input(request.form.get('name', ''), MAX_NAME_LENGTH)
    password = request.form.get('password', '')
    if not username or not email or not name or not password:
        return jsonify({"status": "error", "message": "All fields are required"}), 400
    if '@' not in email:
        return jsonify({"status": "error", "message": "Invalid email address"}), 400
    from core.validation import validate_password
    valid, error = validate_password(password)
    if not valid:
        return jsonify({"status": "error", "message": error}), 400
    try:
        from core.email_utils import send_new_admin_credentials_email, is_smtp_configured
        admin_id = add_admin_with_email(username, email, name, password)
        email_sent = False
        if is_smtp_configured():
            try:
                login_url = url_for('auth.login_page', _external=True)
                email_sent = send_new_admin_credentials_email(
                    email=email,
                    name=name,
                    username=username,
                    password=password,
                    login_url=login_url
                )
            except Exception as e:
                
                log_suspicious_activity('admin_creation_email_failed', {
                    'error': str(e),
                    'admin_id': admin_id
                })
        log_suspicious_activity('admin_created', {
            'created_by': session['admin_id'],
            'new_admin_id': admin_id,
            'new_admin_username': username
        })
        if email_sent:
            message = f"Admin account for {name} created successfully. Login credentials have been sent to {email}."
        else:
            message = f"Admin account for {name} created successfully. ⚠️ Email notification could not be sent - please share credentials manually."
        
        return jsonify({
            "status": "success",
            "message": message,
            "email_sent": email_sent
        })
        
    except ValueError as ve:
        return jsonify({"status": "error", "message": str(ve)}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        log_suspicious_activity('admin_creation_error', {
            'error': str(e),
            'attempted_by': session['admin_id']
        })
        return jsonify({"status": "error", "message": "Failed to create admin account"}), 500

@admin_bp.route('/admin-list', methods=['GET'])
@admin_login_required
def get_admin_list():
    """Get list of all admins (only first admin can see this)"""
    if 'admin_id' not in session:
        return jsonify({"status": "error", "message": "Not authorized"}), 401
    from core.database import get_all_admins, is_first_admin
    if not is_first_admin(session['admin_id']):
        return jsonify({
            "status": "error", 
            "message": "Access denied"
        }), 403
    
    try:
        admins = get_all_admins()
        
        return jsonify({
            "status": "success",
            "admins": [
                {
                    "admin_id": a[0],
                    "username": a[1],
                    "email": a[2],
                    "name": a[3],
                    "created_at": a[4],
                    "email_verified": bool(a[5])
                } for a in admins
            ]
        })
    except Exception as e:
        log_suspicious_activity('get_admin_list_error', {'error': str(e)})
        return jsonify({"status": "error", "message": "Failed to retrieve admin list"}), 500
    
@admin_bp.route('/delete-admin/<int:admin_id>', methods=['POST'])
@csrf_protect
@admin_login_required
def delete_admin_route(admin_id):
    """Delete admin account (only first admin can do this)"""
    if 'admin_id' not in session:
        return jsonify({"status": "error", "message": "Not authorized"}), 401
    
    from core.database import is_first_admin, delete_admin

    if not is_first_admin(session['admin_id']):
        log_suspicious_activity('unauthorized_admin_deletion', {
            'attempted_by': session['admin_id'],
            'target_admin_id': admin_id
        })
        return jsonify({
            "status": "error", 
            "message": "Only the first administrator can delete admin accounts."
        }), 403
        
    if admin_id == session['admin_id']:
        return jsonify({
            "status": "error",
            "message": "You cannot delete your own admin account."
        }), 400
    
    if admin_id == 1:
        return jsonify({
            "status": "error",
            "message": "The first administrator account cannot be deleted for security reasons."
        }), 400
    
    try:
        delete_admin(admin_id)
        
        log_suspicious_activity('admin_deleted', {
            'deleted_by': session['admin_id'],
            'deleted_admin_id': admin_id
        })
        
        return jsonify({
            "status": "success",
            "message": "Admin account deleted successfully"
        })
        
    except ValueError as ve:
        return jsonify({"status": "error", "message": str(ve)}), 404
    except Exception as e:
        import traceback
        traceback.print_exc()
        log_suspicious_activity('admin_deletion_error', {
            'error': str(e),
            'attempted_by': session['admin_id'],
            'target_admin_id': admin_id
        })
        return jsonify({"status": "error", "message": "Failed to delete admin account"}), 500
    
@admin_bp.route('/edit-user/<int:user_id>', methods=['POST'])
@csrf_protect
@admin_login_required
def edit_user_route(user_id):
    """Admin edits an existing user"""
    if 'admin_id' not in session:
        return jsonify({"status": "error", "message": "Not authorized"}), 401
    
    try:
        from core.database import get_user_by_id, get_user_email
        
        # Get current user data
        user = get_user_by_id(user_id)
        if not user:
            return jsonify({"status": "error", "message": "User not found"}), 404
        
        user_id_db, username, old_name, old_role, qr_code, qr_token, status = user
        old_email = get_user_email(user_id)
        
        # Get new data from form
        new_email = sanitize_input(request.form.get('email', ''), MAX_EMAIL_LENGTH).lower()
        new_name = sanitize_input(request.form.get('name', ''), MAX_NAME_LENGTH)
        new_role = sanitize_input(request.form.get('role', ''), 50)
        
        # Validation
        if not new_email or not new_name or not new_role:
            return jsonify({"status": "error", "message": "All fields are required"}), 400
        
        if '@' not in new_email:
            return jsonify({"status": "error", "message": "Invalid email address"}), 400
        
        if new_name.strip() == '':
            return jsonify({"status": "error", "message": "Name cannot be empty"}), 400
        
        if new_role not in ['Staff', 'Employee', 'Student', 'Teacher', 'Visitor']:
            return jsonify({"status": "error", "message": "Invalid role"}), 400
        
        # Track what changed
        changes = {}
        if old_email != new_email:
            changes['email'] = {'old': old_email or 'No email', 'new': new_email}
        if old_name != new_name:
            changes['name'] = {'old': old_name, 'new': new_name}
        if old_role != new_role:
            changes['role'] = {'old': old_role, 'new': new_role}
        
        # If nothing changed
        if not changes:
            return jsonify({"status": "info", "message": "No changes detected"}), 200
        
        # Update database
        conn = get_conn()
        try:
            cur = conn.cursor()
            
            # Update user details
            cur.execute("""
                UPDATE users 
                SET name = ?, role = ?
                WHERE user_id = ?
            """, (new_name, new_role, user_id))
            
            # Update email if changed
            if 'email' in changes:
                # Check if user has an email record
                cur.execute("SELECT COUNT(*) FROM user_emails WHERE user_id = ?", (user_id,))
                has_email = cur.fetchone()[0] > 0
                
                if has_email:
                    cur.execute("""
                        UPDATE user_emails 
                        SET email = ? 
                        WHERE user_id = ?
                    """, (new_email, user_id))
                else:
                    # Insert new email record
                    cur.execute("""
                        INSERT INTO user_emails (user_id, email, email_verified)
                        VALUES (?, ?, 0)
                    """, (user_id, new_email))
            
            conn.commit()
        finally:
            conn.close()
        
        # Send notification email
        from core.email_utils import send_user_profile_updated_email, is_smtp_configured
        
        email_sent = False
        if is_smtp_configured() and new_email:
            try:
                email_sent = send_user_profile_updated_email(
                    email=new_email,
                    name=new_name,
                    username=username,
                    changes=changes,
                    updated_by=session.get('admin_name', session.get('admin_username', 'Administrator'))
                )
            except Exception as e:
                log_suspicious_activity('user_update_email_failed', {
                    'error': str(e),
                    'user_id': user_id
                })
        
        try:
            import routes_sse
            routes_sse.notify_user_profile_changed(user_id, {
                'name': new_name,
                'role': new_role,
                'email': new_email
            })
            print(f"[Admin] ✅ SSE notification sent to user {user_id}")
        except ImportError as ie:
            print(f"[Admin] ⚠️ SSE module not available: {ie}")
        except AttributeError as ae:
            print(f"[Admin] ⚠️ SSE function not found: {ae}")
        except Exception as e:
            print(f"[Admin] ⚠️ Failed to send SSE notification: {e}")

        
        # Log the update
        log_suspicious_activity('user_profile_updated', {
            'admin_id': session['admin_id'],
            'admin_username': session.get('admin_username'),
            'target_user_id': user_id,
            'target_username': username,
            'changes': changes,
            'email_sent': email_sent
        })
        
        # Build response message
        change_summary = []
        for field, vals in changes.items():
            change_summary.append(f"{field.capitalize()}: {vals['old']} → {vals['new']}")
        
        message = f"User '{new_name}' updated successfully.<br>Changes: " + ", ".join(change_summary)
        
        if email_sent:
            message += f"<br>Notification sent to {new_email}"
        elif is_smtp_configured():
            message += "<br>⚠️ Failed to send email notification"
        
        return jsonify({
            "status": "success",
            "message": message,
            "changes": changes
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        log_suspicious_activity('edit_user_error', {
            'admin_id': session.get('admin_id'),
            'error': str(e),
            'user_id': user_id
        })
        return jsonify({
            "status": "error",
            "message": "Failed to update user. Please try again."
        }), 500
